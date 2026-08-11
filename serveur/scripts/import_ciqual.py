# scripts/import_ciqual.py
#
# Import one-shot de la table CIQUAL (ANSES, ~3 200 aliments generiques FR)
# dans le catalogue `foods` du user starter_template (Nutrition V2, decision D9
# de docs/NUTRITION_DESIGN.md : catalogue user-scoped copie au /signup via
# copy_starter_pack, comme les muscles/exercises).
#
# Source : https://ciqual.anses.fr/ -> "Telecharger la table" (XLS/CSV).
# Le fichier .xls historique (BIFF) n'est PAS lisible directement : le
# convertir en .xlsx (Excel/LibreOffice) ou exporter en .csv (separateur `;`,
# decimales a virgule — defaults FR).
#
# Usage (depuis serveur/, venv active) :
#     python scripts/import_ciqual.py "C:\chemin\Table Ciqual 2020.csv"
#     python scripts/import_ciqual.py "C:\chemin\Table Ciqual 2020.xlsx"
#
# Idempotent : dedup par (user_id=template, source='CIQUAL', source_ref=alim_code).
# Re-run = update des macros des rows existantes + insert des nouvelles.
#
# ⚠️ `fill_database.py` droppe la table foods (clear all sauf users) : relancer
# cet import APRES un re-seed si on veut le catalogue CIQUAL en dev.
# Les users deja crees ne recoivent rien automatiquement (le copy ne se fait
# qu'au /signup) : lancer `python scripts/backfill_nutrition.py --all` apres
# cet import pour les servir.

from __future__ import annotations

import asyncio
import sys
import unicodedata
from pathlib import Path

# Lancable depuis serveur/ (python scripts/import_ciqual.py) : ajoute serveur/
# au path pour importer app.*.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

# Module pur (pas de DB) : mapping libelles CIQUAL -> groupe curate.
from app.food_taxonomy import classify_ciqual  # noqa: E402


# ============================================================================
# Parsing pur (testable sans DB)
# ============================================================================

def _normalize_header(header: str) -> str:
    """Minuscule + accents retires, pour le matching tolerant des colonnes."""
    nfkd = unicodedata.normalize("NFKD", header or "")
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()


def parse_nutrient(raw) -> float | None:
    """Nettoie une valeur nutritionnelle CIQUAL.

    Conventions ANSES : decimales a virgule, 'traces' (= negligeable -> 0.0),
    '< x' (sous le seuil de detection -> borne haute x, choix conservateur),
    '-' ou vide (= non determine -> None).
    """
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s or s == "-":
        return None
    if _normalize_header(s) == "traces":
        return 0.0
    s = s.lstrip("<").strip().replace(",", ".").replace(" ", "")
    try:
        return float(s)
    except ValueError:
        return None


# Matching des colonnes CIQUAL : pour chaque champ, liste de candidats par
# priorite ; un candidat = tuple de substrings qui doivent TOUTES apparaitre
# dans le header normalise. La table 2020 a 2 colonnes energie kcal (Reglement
# UE 1169/2011 + Jones) et des jumelles en kJ -> on exige "kcal" ; 2 colonnes
# proteines (Jones + brutes N x 6.25) -> on prefere Jones.
_COLUMN_CANDIDATES: dict[str, list[tuple[str, ...]]] = {
    "code": [("alim_code",)],
    "name": [("alim_nom_fr",)],
    "kcal": [("energie", "reglement ue", "kcal"), ("energie", "kcal"), ("kcal",)],
    "protein": [("proteines", "jones"), ("proteines",)],
    "carbs": [("glucides",)],
    "fat": [("lipides",)],
    # ⚠️ Exiger "fibres alimentaires" : un simple "fibres" matcherait la colonne
    # "Energie, N x facteur Jones, avec fibres (kJ/100 g)" (placee AVANT la vraie
    # colonne "Fibres alimentaires") -> on stockerait l'energie en kJ. Bug 2026-06-13.
    "fiber": [("fibres", "alimentaires")],
    # Categorisation (feature Categories d'aliments) : groupe + sous-groupe CIQUAL
    # (deja dans la table) -> mappes vers un groupe curate via food_taxonomy.
    # Optionnels (anciens exports sans ces colonnes : food_group reste None).
    # alim_ssssgrp (3e niveau) tranche viande rouge/blanche (poulet/dinde/boeuf...) :
    # le 2e niveau du CIQUAL 2025 reel est trop grossier (viandes crues/cuites).
    "alim_grp": [("alim_grp_nom_fr",)],
    "alim_ssgrp": [("alim_ssgrp_nom_fr",)],
    "alim_ssssgrp": [("alim_ssssgrp_nom_fr",)],
    "sugar": [("sucres",)],
    "sat_fat": [("ag satures",)],
    # "sel" seul matcherait "selenium" -> on exige le libelle complet d'abord.
    "salt": [("sel chlorure",), ("sel (",)],
    # Vitamines & mineraux (pack essentiel ~10, D11 etendu 2026-06-13). Matching
    # strict pour eviter les sous-chaines ambigues (cf. bug fibres) :
    #  - "sodium" seul matcherait "Sel chlorure de sodium" (en g) -> exiger "mg".
    #  - "fer" seul matcherait "facteur" -> exiger "fer (" en mg.
    #  - vitamine A n'a PAS de colonne CIQUAL : RAE = retinol + beta-carotene/12,
    #    derive de deux colonnes (cf. extract_food_specs).
    "calcium": [("calcium",)],
    "iron": [("fer", "mg")],
    "magnesium": [("magnesium",)],
    "potassium": [("potassium",)],
    "sodium": [("sodium", "mg")],
    "zinc": [("zinc",)],
    "retinol": [("retinol",)],
    "beta_carotene": [("carotene",)],
    "vitamin_d": [("vitamine d",)],
    "vitamin_c": [("vitamine c",)],
    "vitamin_b12": [("vitamine b12",)],
}


def _resolve_columns(headers: list[str]) -> dict[str, int]:
    """Mappe {champ: index de colonne}. Leve si code/name/kcal introuvables."""
    normalized = [_normalize_header(h) for h in headers]
    mapping: dict[str, int] = {}
    for field, candidates in _COLUMN_CANDIDATES.items():
        for required in candidates:
            idx = next(
                (i for i, h in enumerate(normalized)
                 if all(sub in h for sub in required)),
                None,
            )
            if idx is not None:
                mapping[field] = idx
                break
    missing = [f for f in ("code", "name", "kcal") if f not in mapping]
    if missing:
        raise ValueError(
            f"Colonnes CIQUAL introuvables : {missing}. Headers vus : {headers[:10]}..."
        )
    return mapping


def extract_food_specs(headers: list[str], rows: list[list]) -> list[dict]:
    """Transforme les rows CIQUAL en specs Food normalisees.

    Skip les rows sans code/nom/kcal (un aliment sans energie est inutilisable).
    Macros NOT NULL (protein/carbs/fat) manquantes -> 0.0 ; micro-nutriments
    manquants -> None (colonnes nullable, D11).
    """
    cols = _resolve_columns(headers)

    def cell(row: list, field: str):
        idx = cols.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    specs: list[dict] = []
    for row in rows:
        code = str(cell(row, "code") or "").strip()
        name = str(cell(row, "name") or "").strip()
        kcal = parse_nutrient(cell(row, "kcal"))
        if not code or not name or kcal is None:
            continue
        specs.append({
            "source_ref": code,
            "name": name,
            # Groupe curate derive des libelles CIQUAL groupe/sous-groupe (None si
            # les colonnes alim_grp/alim_ssgrp sont absentes -> classify -> AUTRE).
            # 3e niveau (alim_ssssgrp) optionnel : tranche viande rouge/blanche.
            "food_group": classify_ciqual(
                cell(row, "alim_grp"), cell(row, "alim_ssgrp"), cell(row, "alim_ssssgrp")
            ),
            "kcal_per_100g": kcal,
            "protein_per_100g": parse_nutrient(cell(row, "protein")) or 0.0,
            "carbs_per_100g": parse_nutrient(cell(row, "carbs")) or 0.0,
            "fat_per_100g": parse_nutrient(cell(row, "fat")) or 0.0,
            "fiber_per_100g": parse_nutrient(cell(row, "fiber")),
            "sugar_per_100g": parse_nutrient(cell(row, "sugar")),
            "sat_fat_per_100g": parse_nutrient(cell(row, "sat_fat")),
            "salt_per_100g": parse_nutrient(cell(row, "salt")),
            # Vitamines & mineraux (pack essentiel ~10, D11 etendu).
            "iron_per_100g": parse_nutrient(cell(row, "iron")),
            "calcium_per_100g": parse_nutrient(cell(row, "calcium")),
            "magnesium_per_100g": parse_nutrient(cell(row, "magnesium")),
            "zinc_per_100g": parse_nutrient(cell(row, "zinc")),
            "potassium_per_100g": parse_nutrient(cell(row, "potassium")),
            "sodium_per_100g": parse_nutrient(cell(row, "sodium")),
            "vitamin_c_per_100g": parse_nutrient(cell(row, "vitamin_c")),
            "vitamin_d_per_100g": parse_nutrient(cell(row, "vitamin_d")),
            "vitamin_b12_per_100g": parse_nutrient(cell(row, "vitamin_b12")),
            # Vitamine A = RAE (µg) = retinol + beta-carotene/12 (CIQUAL n'a pas de
            # colonne 'vit A' unique). None seulement si les DEUX sources manquent.
            "vitamin_a_per_100g": _vitamin_a_rae(
                parse_nutrient(cell(row, "retinol")),
                parse_nutrient(cell(row, "beta_carotene")),
            ),
        })
    return specs


def _vitamin_a_rae(retinol_ug: float | None, beta_carotene_ug: float | None) -> float | None:
    """Vitamine A en µg RAE : retinol + beta-carotene/12 (conversion ANSES/IOM).

    Retourne None si retinol ET beta-carotene sont indetermines (pas de donnee) ;
    sinon traite la source manquante comme 0.0 (apport nul de cette voie).
    """
    if retinol_ug is None and beta_carotene_ug is None:
        return None
    return (retinol_ug or 0.0) + (beta_carotene_ug or 0.0) / 12.0


# ============================================================================
# Lecture fichier (CSV `;` ou XLSX)
# ============================================================================

def read_table(path: Path) -> tuple[list[str], list[list]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        import csv
        for encoding in ("utf-8-sig", "latin-1"):
            try:
                with path.open(newline="", encoding=encoding) as f:
                    reader = csv.reader(f, delimiter=";")
                    rows = list(reader)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError(f"Encodage illisible : {path}")
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise SystemExit(
                "openpyxl requis pour lire un .xlsx : `pip install openpyxl`, "
                "ou exporter le fichier en .csv (separateur ;)."
            )
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
    elif suffix == ".xls":
        raise SystemExit(
            "Format .xls (BIFF) non supporte : convertir en .xlsx ou .csv "
            "(Excel/LibreOffice : Enregistrer sous)."
        )
    else:
        raise SystemExit(f"Extension non supportee : {suffix} (attendu .csv ou .xlsx)")

    if not rows:
        raise ValueError(f"Fichier vide : {path}")
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return headers, rows[1:]


# ============================================================================
# Upsert DB (starter_template)
# ============================================================================

async def import_into_template(specs: list[dict]) -> tuple[int, int]:
    """Upsert les specs dans foods du user starter_template.

    Dedup par source_ref (source='CIQUAL'). Retourne (inserted, updated).
    """
    import uuid as uuidlib

    from sqlalchemy import select

    from app.database import AsyncSessionLocal
    from app.models.food import Food
    from app.models.user import User
    from app.settings import settings

    async with AsyncSessionLocal() as db:
        template_id = (await db.execute(
            select(User.id).where(User.username == settings.STARTER_TEMPLATE_USERNAME)
        )).scalar_one_or_none()
        if template_id is None:
            raise SystemExit(
                f"User '{settings.STARTER_TEMPLATE_USERNAME}' introuvable. "
                "Lancer `python -m app.fill_database` (dev) pour le seeder."
            )

        existing_rows = (await db.execute(
            select(Food).where(Food.user_id == template_id, Food.source == "CIQUAL")
        )).scalars().all()
        existing_by_ref = {f.source_ref: f for f in existing_rows}

        inserted = updated = 0
        nutrient_fields = (
            "name", "food_group", "kcal_per_100g", "protein_per_100g", "carbs_per_100g",
            "fat_per_100g", "fiber_per_100g", "sugar_per_100g",
            "sat_fat_per_100g", "salt_per_100g",
            "iron_per_100g", "calcium_per_100g", "magnesium_per_100g",
            "zinc_per_100g", "potassium_per_100g", "sodium_per_100g",
            "vitamin_c_per_100g", "vitamin_d_per_100g", "vitamin_b12_per_100g",
            "vitamin_a_per_100g",
        )
        for spec in specs:
            current = existing_by_ref.get(spec["source_ref"])
            if current is not None:
                for field in nutrient_fields:
                    setattr(current, field, spec[field])
                updated += 1
            else:
                db.add(Food(
                    uuid=str(uuidlib.uuid4()),
                    user_id=template_id,
                    source="CIQUAL",
                    source_ref=spec["source_ref"],
                    name=spec["name"],
                    brand=None,
                    food_group=spec["food_group"],
                    kcal_per_100g=spec["kcal_per_100g"],
                    protein_per_100g=spec["protein_per_100g"],
                    carbs_per_100g=spec["carbs_per_100g"],
                    fat_per_100g=spec["fat_per_100g"],
                    fiber_per_100g=spec["fiber_per_100g"],
                    sugar_per_100g=spec["sugar_per_100g"],
                    sat_fat_per_100g=spec["sat_fat_per_100g"],
                    salt_per_100g=spec["salt_per_100g"],
                    iron_per_100g=spec["iron_per_100g"],
                    calcium_per_100g=spec["calcium_per_100g"],
                    magnesium_per_100g=spec["magnesium_per_100g"],
                    zinc_per_100g=spec["zinc_per_100g"],
                    potassium_per_100g=spec["potassium_per_100g"],
                    sodium_per_100g=spec["sodium_per_100g"],
                    vitamin_c_per_100g=spec["vitamin_c_per_100g"],
                    vitamin_d_per_100g=spec["vitamin_d_per_100g"],
                    vitamin_b12_per_100g=spec["vitamin_b12_per_100g"],
                    vitamin_a_per_100g=spec["vitamin_a_per_100g"],
                    is_favorite=False,
                    archived=False,
                ))
                inserted += 1
        await db.commit()
    return inserted, updated


async def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage : python scripts/import_ciqual.py <table_ciqual.csv|.xlsx>")
    path = Path(sys.argv[1])
    if not path.is_file():
        raise SystemExit(f"Fichier introuvable : {path}")

    headers, rows = read_table(path)
    specs = extract_food_specs(headers, rows)
    print(f"📄 {len(rows)} rows lues, {len(specs)} aliments exploitables (code+nom+kcal).")

    inserted, updated = await import_into_template(specs)
    print(f"✅ Import CIQUAL termine : {inserted} inserted, {updated} updated "
          "(catalogue starter_template — copie aux nouveaux users au /signup).")


if __name__ == "__main__":
    asyncio.run(main())
